import requests
from openpyxl import load_workbook
from datetime import datetime, timezone
import pytz
import time
from colorama import init, Fore
import random
import json  # Tambahkan import json

# Inisialisasi colorama
init(autoreset=True)

def random_color():
    # Pilih warna acak dari Fore, kecuali Fore.RED
    colors = [Fore.GREEN, Fore.YELLOW, Fore.BLUE, Fore.MAGENTA, Fore.CYAN, Fore.WHITE]
    return random.choice(colors)

def countdown(seconds):
    """Menampilkan hitung mundur dalam detik dengan membersihkan baris sebelumnya."""
    for remaining in range(seconds, 0, -1):
        # Buat string placeholder dengan panjang yang cukup untuk menghapus teks lama
        placeholder = " " * 60
        # Tampilkan hitung mundur
        print(f"\rCountdown {remaining} . . . . ." + placeholder, end="")
        time.sleep(1)
    print("\r" + " " * 60, end="\n")  # Bersihkan baris hitung mundur

def main(file_path):
    def get_latest_price_by_id(id, convert_id='2781'):
        base_url = "https://pro-api.coinmarketcap.com/v1/tools/price-conversion"
        # Baca API key dari apikey.json
        with open('apikey.json') as f:
            apikeys = json.load(f)
        apikey = apikeys.get("coinmarketcap", "")

        params = {
            'id': id,
            'amount': 1,
            'convert_id': convert_id
        }
        
        headers = {
            'Accepts': 'application/json',
            'X-CMC_PRO_API_KEY': apikey,
        }
        
        try:
            print(f"{random_color()}(CMC) Get Data ID For: {id}")
            response = requests.get(base_url, params=params, headers=headers)
            data = response.json()
            if response.status_code != 200:
                print(f"{Fore.RED}API Error {response.status_code}: {response.text}")
                return None
            
            if 'data' in data and data['data']:
                quote_data = data['data']['quote'][convert_id]
                price = quote_data['price']
                last_updated_iso = quote_data['last_updated']
                
                # Convert ISO 8601 string to datetime object
                last_updated_dt = datetime.fromisoformat(last_updated_iso.replace('Z', '+00:00'))
                # Convert datetime object to timestamp (seconds since epoch)
                last_updated_timestamp = int(last_updated_dt.replace(tzinfo=timezone.utc).timestamp())
                
                return {
                    'price': price,
                    'last_updated': last_updated_timestamp
                }
            else:
                # Menambahkan padding untuk membuat teks menjorok
                print(f"{Fore.RED}Already Updated ID: {id}")
                return None
        
        except Exception as e:
            # Menambahkan padding untuk membuat teks menjorok
            print(f"{Fore.RED}Error: {str(e)}")
            return None

    sheet_name = "KONVERSI"

    try:
        # Load the workbook
        wb = load_workbook(file_path)
        sheet = wb[sheet_name]
    except FileNotFoundError:
        # Menambahkan padding untuk membuat teks menjorok
        print(f"{Fore.RED}Excel File '{file_path}' Not Found.")
        return
    except Exception as e:
        # Menambahkan padding untuk membuat teks menjorok
        print(f"{Fore.RED}Error loading workbook: {str(e)}")
        return

    # Initialize empty list to store IDs from Excel
    ids_to_check = []

    # Iterate over IDs in Excel sheet column A (assuming IDs are in column A starting from row 2)
    for row in sheet.iter_rows(min_row=2, max_col=1, max_row=sheet.max_row, values_only=True):
        id_value = row[0]
        if id_value is not None and id_value != "N/A":
            ids_to_check.append(str(id_value))
        else:
            # Append "N/A" to ids_to_check to maintain alignment with rows in Excel
            ids_to_check.append("N/A")

    # Function to convert timestamp to WIB format
    def timestamp_to_wib(timestamp):
        tz = pytz.timezone('Asia/Jakarta')  # Set timezone to WIB
        return datetime.fromtimestamp(timestamp, tz).strftime('%Y-%m-%d %H:%M:%S')

    # Track the number of API requests made
    api_request_count = 0

    # Iterate over each ID and get the latest price and last update
    for idx, id in enumerate(ids_to_check, start=2):  # start=2 to skip header row in Excel
        if id != "N/A":
            data = get_latest_price_by_id(id, '2781')  # '2781' is the ID for USD in CoinMarketCap
            api_request_count += 1
            
            if data:
                sheet[f'E{idx}'] = data.get('price', 'N/A')  # CMC PRICE, default to 'N/A' if not found
                
                # Convert last_updated to WIB and write to Excel
                if 'last_updated' in data:
                    sheet[f'H{idx}'] = timestamp_to_wib(data['last_updated'])  # LAST UPDATE (CMC)
                else:
                    sheet[f'H{idx}'] = 'N/A'
        else:
            # For "N/A" ID, write "N/A" to CMC PRICE and LAST UPDATE (CMC)
            sheet[f'E{idx}'] = 'N/A'
            sheet[f'H{idx}'] = 'N/A'
        
        # Check if 10 requests have been made, then wait for 10 seconds
        if api_request_count % 10 == 0:
            countdown(10)

    # Save the workbook
    try:
        wb.save(file_path)
        print(f"{random_color()}Update SUKSES!! Path '{file_path}'\n")
    except Exception as e:
        # Menambahkan padding untuk membuat teks menjorok
        print(f"{Fore.RED}Error saving workbook: {str(e)}")

# Path ke dua file Excel yang berbeda
excel_path_1 = r"C:\My DATA\Gitub Project\price_update\doc\tes.xlsx"
# excel_path_2 = r"C:\My DATA\Gitub Project\price_update\coba.xlsx"