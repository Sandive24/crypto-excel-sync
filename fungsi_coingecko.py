import requests, json
from openpyxl import load_workbook
from datetime import datetime
import pytz
import time
from colorama import init, Fore
import random

init(autoreset=True)

def random_color():
    colors = [Fore.GREEN, Fore.YELLOW, Fore.BLUE, Fore.MAGENTA, Fore.CYAN, Fore.WHITE]
    return random.choice(colors)

def countdown(seconds):
    for remaining in range(seconds, 0, -1):
        placeholder = " " * 60
        print(f"\rCountdown {remaining} Second For Next Requests. . . . ." + placeholder, end="")
        time.sleep(1)
    print("\r" + " " * 60, end="\n")

def run(file_path):
    sheet_name = "KONVERSI"
    try:
        wb = load_workbook(file_path)
        sheet = wb[sheet_name]
    except FileNotFoundError:
        print(f"{Fore.RED}Excel File '{file_path}' Not Found.")
        return
    except Exception as e:
        print(f"{Fore.RED}Error loading workbook: {str(e)}")
        return

    # Ambil ID CoinGecko dari kolom B
    ids_to_check = []
    for row in sheet.iter_rows(min_row=2, min_col=2, max_col=2, max_row=sheet.max_row, values_only=True):
        id_value = row[0]
        if id_value is not None and id_value != "N/A":
            ids_to_check.append(str(id_value))
        else:
            ids_to_check.append("N/A")

    def timestamp_to_wib(timestamp):
        tz = pytz.timezone('Asia/Jakarta')
        return datetime.fromtimestamp(timestamp, tz).strftime('%Y-%m-%d %H:%M:%S')

    # Batch request ke CoinGecko (maks 250 ID per request)
    ids_batch = [id for id in ids_to_check if id != "N/A"]
    if not ids_batch:
        print(f"{Fore.RED}Tidak ada ID CoinGecko yang valid di Excel.")
        return

    ids_str = ",".join(ids_batch)
    url = "https://api.coingecko.com/api/v3/simple/price"
    params = {
        "ids": ids_str,
        "vs_currencies": "usd",
        "include_last_updated_at": "true"
    }

    try:
        response = requests.get(url, params=params)
        data = response.json()
        print("Respon Kode:", response.status_code)

        for idx, coin_id in enumerate(ids_to_check, start=2):
            if coin_id != "N/A":
                print(f"{random_color()}(CG) Get Data ID For: {coin_id}")
                if coin_id in data:
                    price = data[coin_id].get("usd", "N/A")
                    last_updated_unix = data[coin_id].get("last_updated_at", None)
                    if last_updated_unix:
                        last_updated_str = timestamp_to_wib(last_updated_unix)
                    else:
                        last_updated_str = "N/A"
                    sheet[f'F{idx}'] = price
                    sheet[f'I{idx}'] = last_updated_str
                else:
                    sheet[f'F{idx}'] = "N/A"
                    sheet[f'I{idx}'] = "N/A"
            else:
                sheet[f'F{idx}'] = "N/A"
                sheet[f'I{idx}'] = "N/A"

        wb.save(file_path)
        print(f"{random_color()}(CG) Update SUKSES!! Path '{file_path}'\n")
    except Exception as e:
        print(f"{Fore.RED}Error saat request ke CoinGecko: {str(e)}")

# Daftar file yang mau di Automatisasi:
excel_path_1 = r"C:\My DATA\Gitub Project\price_update\doc\tes.xlsx"
# excel_path_2 = r"C:\My DATA\Gitub Project\price_update\coba.xlsx"