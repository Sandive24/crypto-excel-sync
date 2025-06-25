from PIL import Image
import pytesseract
import re
import pandas as pd
import os
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font
from datetime import datetime

def update_excel(excel_path, sheet_name, df):
    # Membuka file Excel yang ada
    wb = load_workbook(excel_path)

    # Periksa apakah sheet yang diperlukan ada
    if sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
    else:
        ws = wb.create_sheet(title=sheet_name)

    # Membaca data yang sudah ada di sheet
    existing_data = pd.DataFrame(ws.values)

    # Mengatur header jika sheet kosong
    if existing_data.empty:
        ws.append(["NAMA", "VOLUME", "VOLUME UNIT", "REWARD", "REWARD UNIT", "MAX REWARD", "MAX REWARD UNIT"])
        existing_data = pd.DataFrame(columns=["NAMA", "VOLUME", "VOLUME UNIT", "REWARD", "REWARD UNIT", "MAX REWARD", "MAX REWARD UNIT"])

    # Update data di sheet
    for idx, row_data in df.iterrows():
        name = row_data["NAMA"]
        volume = row_data["VOLUME"]
        volume_unit = row_data["VOLUME UNIT"]
        reward = row_data["REWARD"]
        reward_unit = row_data["REWARD UNIT"]
        max_reward = row_data["MAX REWARD"]
        max_reward_unit = row_data["MAX REWARD UNIT"]

        # Cek apakah data dengan nama yang sama sudah ada di sheet
        found = False
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=7):
            if row[0].value == name:
                # Update data yang ada
                row[1].value = volume
                row[2].value = volume_unit
                row[3].value = reward
                row[4].value = reward_unit
                row[5].value = max_reward
                row[6].value = max_reward_unit
                found = True
                break

        # Jika data tidak ditemukan, tambahkan sebagai baris baru
        if not found:
            ws.append([name, volume, volume_unit, reward, reward_unit, max_reward, max_reward_unit])

    # Atur lebar kolom dan alignment untuk kolom A-G
    for col in ws.columns:
        col_letter = col[0].column_letter
        ws.column_dimensions[col_letter].width = 25
        for cell in col:
            cell.alignment = Alignment(horizontal='center', vertical='center')

    # Mengatur format angka untuk kolom B, D, dan F
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        row[1].number_format = '#,##0'  # VOLUME
        row[3].number_format = '#,##0.000'  # REWARD
        row[5].number_format = '#,##0.00'  # MAX REWARD

    # Atur kolom H untuk "Last Updated"
    ws.column_dimensions['H'].width = 35

    # Menambahkan informasi "Last Updated" di H1
    ws['H1'] = "Last Updated"
    ws['H1'].alignment = Alignment(horizontal='center', vertical='center')
    ws['H1'].font = Font(bold=True)

    # Menambahkan informasi tanggal dan waktu terakhir diperbarui di H2
    ws['H2'] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    ws['H2'].alignment = Alignment(horizontal='center', vertical='center')

    # Menyimpan workbook yang telah diperbarui
    wb.save(excel_path)

def main():
    # Menentukan path ke executable Tesseract secara eksplisit
    pytesseract.pytesseract.tesseract_cmd = r'C:\Program Files\Tesseract-OCR\tesseract.exe'

    # Path ke folder gambar
    folder_path = r"C:\Users\iksan\OneDrive\img"

    # Path ke file Excel yang akan digunakan
    excel_paths = [
        r"C:\My DATA\Gitub Project\price_update\doc\tes.xlsx"
        # r"C:\My DATA\Gitub Project\price_update\coba.xlsx"
    ]
    sheet_name = 'REWARD'

    # Daftar untuk menyimpan data ekstraksi
    data = []

    # Mengambil semua file gambar dari folder
    for file_name in os.listdir(folder_path):
        # Hanya memproses file dengan ekstensi gambar yang umum
        if file_name.lower().endswith(('.png', '.jpg', '.jpeg')):
            try:
                # Path lengkap ke file gambar
                image_path = os.path.join(folder_path, file_name)
                
                # Membuka file gambar
                image = Image.open(image_path)
                
                # Menggunakan pytesseract untuk mengubah gambar menjadi teks
                extracted_text = pytesseract.image_to_string(image, lang='eng')
                
                # Menghilangkan simbol dan karakter non-alfanumerik kecuali spasi, koma, dan titik
                cleaned_text = re.sub(r'[^\w\s.,/]', '', extracted_text)

                # Ekstraksi informasi dari teks yang dibersihkan
                name_match = re.search(r'worth\s+of\s+(\w+)', cleaned_text)
                volume_match = re.search(r'Trade at least\s*([\d\.,/]+)\s*(USDT|ETH|A8|FET|UXLINK|)', cleaned_text)
                reward_match = re.search(r'Your Estimated Rewards\s*([\d\.,]+)\s*(USDT|ETH|A8|FET|UXLINK|)', cleaned_text)
                max_reward_match = re.search(r'Earn Up to\s*([\d\.,]+)\s*(USDT|ETH|A8|FET|UXLINK|)', cleaned_text)

                # Mengambil data dari hasil ekstraksi
                name = name_match.group(1) if name_match else "Tidak Ditemukan"

                # Memproses volume
                volume = volume_match.group(1) if volume_match else "Tidak Ditemukan"
                volume_unit = volume_match.group(2) if volume_match else "Tidak Ditemukan"
                volume_cleaned = volume.split('/')[0].replace('>', '').strip()  # Mengambil bagian sebelum '/'
                
                # Format volume dengan pemisah ribuan koma
                volume_number = re.sub(r'[^\d]', '', volume_cleaned)  # Hapus simbol selain digit
                if volume_number:
                    volume_formatted = int(volume_number)  # Simpan sebagai angka
                else:
                    volume_formatted = volume_cleaned
                
                volume_with_unit = f"{volume_formatted} {volume_unit}"

                # Memproses reward dan max reward
                reward = reward_match.group(1) if reward_match else "Tidak Ditemukan"
                reward_unit = reward_match.group(2) if reward_match else "Tidak Ditemukan"

                max_reward = max_reward_match.group(1) if max_reward_match else ""
                max_reward_unit = max_reward_match.group(2) if max_reward_match.group(2) else ""

                # Mengkonversi reward dan max reward menjadi angka jika memungkinkan
                reward_number = float(reward.replace(',', '')) if reward.replace(',', '').replace('.', '').isdigit() else reward
                max_reward_number = float(max_reward.replace(',', '')) if max_reward.replace(',', '').replace('.', '').isdigit() else max_reward

                # Menyertakan satuan ke dalam hasil ekstraksi
                reward_with_unit = f"{reward_number} {reward_unit}"
                max_reward_with_unit = f"{max_reward_number} {max_reward_unit}"

                # Menambahkan data ke daftar
                data.append([name, volume_formatted, volume_unit, reward_number, reward_unit, max_reward_number, max_reward_unit])
            
            except Exception as e:
                print(f"Error processing file {file_name}: {e}")

    # Membuat DataFrame dari data
    df = pd.DataFrame(data, columns=["NAMA", "VOLUME", "VOLUME UNIT", "REWARD", "REWARD UNIT", "MAX REWARD", "MAX REWARD UNIT"])

    # Update kedua file Excel
    for excel_path in excel_paths:
        update_excel(excel_path, sheet_name, df)
        print(f"Done Update To {excel_path}")

if __name__ == "__main__":
    main()
